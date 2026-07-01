import argparse
import json
import math
import os
import re
import time
import unicodedata
from collections import defaultdict
from datetime import datetime

import requests
from openpyxl import load_workbook


def normalize_text(value):
    if value is None:
        return ""
    s = str(value).strip()
    s = " ".join(s.split())
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.lower()


def to_number(value, default=0.0):
    if value is None or value == "":
        return float(default)
    try:
        if isinstance(value, str):
            value = value.replace(",", "").strip()
        n = float(value)
        if math.isnan(n) or math.isinf(n):
            return float(default)
        return n
    except Exception:
        return float(default)


def infer_tipo(tipo_raw):
    t = normalize_text(tipo_raw)
    if "equipo" in t:
        return "equipo"
    if "herramienta" in t:
        return "herramienta"
    if "material" in t:
        return "material"
    return "general"


def parse_throttle_wait(error_text, default_seconds=20):
    match = re.search(r"Expected available in\s+(\d+)\s+seconds", error_text)
    if match:
        return max(5, int(match.group(1)) + 2)
    return default_seconds


class ApiClient:
    def __init__(self, base_url, sso_token):
        self.base_url = base_url.rstrip("/")
        self.sso_token = sso_token
        self.token = None

    def login(self):
        url = f"{self.base_url}/users/sso-login/"
        resp = requests.post(url, json={"sso_token": self.sso_token}, timeout=30)
        resp.raise_for_status()
        payload = resp.json()
        data = payload.get("data") or {}
        token = data.get("access_token")
        if not token:
            raise RuntimeError(f"No se obtuvo access_token: {payload}")
        self.token = token

    def request(self, method, path, retries=8, **kwargs):
        if not self.token:
            raise RuntimeError("Token no inicializado. Ejecuta login primero.")

        url = f"{self.base_url}/{path.lstrip('/')}"
        headers = kwargs.pop("headers", {})
        headers["Authorization"] = f"Bearer {self.token}"
        headers.setdefault("Content-Type", "application/json")

        attempt = 0
        while attempt < retries:
            attempt += 1
            resp = requests.request(method, url, headers=headers, timeout=35, **kwargs)
            try:
                payload = resp.json()
            except Exception:
                payload = {"raw": resp.text}

            if resp.status_code == 429 and attempt < retries:
                wait_s = parse_throttle_wait(str(payload), 20)
                time.sleep(wait_s)
                continue

            if resp.status_code >= 400:
                raise RuntimeError(f"{method} {path} -> {resp.status_code}: {payload}")

            if isinstance(payload, dict) and payload.get("success") is False:
                msg = str(payload)
                if "throttled" in msg.lower() and attempt < retries:
                    wait_s = parse_throttle_wait(msg, 20)
                    time.sleep(wait_s)
                    continue
                raise RuntimeError(f"{method} {path} -> API error: {payload}")

            if isinstance(payload, dict) and "data" in payload:
                return payload["data"]
            return payload

        raise RuntimeError(f"{method} {path} -> agotados reintentos")


def collect_excel_data(excel_path):
    wb = load_workbook(excel_path, keep_vba=True, data_only=False)
    ws_ing = wb["INGRESOS"]
    ws_egr = wb["EGRESOS"]

    catalog = {}
    ingresos_qty = defaultdict(float)
    egresos_qty = defaultdict(float)
    dispatch_rows = []

    def capture_catalog(code, desc, tipo):
        if not code:
            return
        if code not in catalog:
            catalog[code] = {
                "codigo": code,
                "nombre": desc or code,
                "tipo_item": infer_tipo(tipo),
            }

    for r in range(2, ws_ing.max_row + 1):
        code = ws_ing.cell(r, 1).value
        desc = ws_ing.cell(r, 2).value
        tipo = ws_ing.cell(r, 3).value
        qty = ws_ing.cell(r, 7).value

        if code is None:
            continue
        code = str(code).strip()
        if not code:
            continue

        desc = str(desc).strip() if desc is not None else ""
        capture_catalog(code, desc, tipo)
        ingresos_qty[code] += to_number(qty, 0)

    for r in range(2, ws_egr.max_row + 1):
        code = ws_egr.cell(r, 1).value
        desc = ws_egr.cell(r, 2).value
        tipo = ws_egr.cell(r, 3).value
        cliente = ws_egr.cell(r, 4).value
        documento = ws_egr.cell(r, 5).value
        fecha = ws_egr.cell(r, 6).value
        qty = ws_egr.cell(r, 7).value

        if code is None:
            continue
        code = str(code).strip()
        if not code:
            continue

        desc = str(desc).strip() if desc is not None else ""
        capture_catalog(code, desc, tipo)
        egresos_qty[code] += to_number(qty, 0)

        dispatch_rows.append(
            {
                "codigo": code,
                "descripcion": desc,
                "tipo": str(tipo).strip() if tipo is not None else "",
                "cliente": str(cliente).strip() if cliente is not None else "",
                "documento": str(documento).strip() if documento is not None else "",
                "fecha": str(fecha) if fecha is not None else "",
                "cantidad": to_number(qty, 0),
            }
        )

    items = []
    for code, meta in catalog.items():
        saldo = int(round(max(0.0, ingresos_qty.get(code, 0.0) - egresos_qty.get(code, 0.0))))
        items.append(
            {
                "codigo": code,
                "nombre": meta["nombre"],
                "tipo_item": meta["tipo_item"],
                "cantidad": saldo,
            }
        )

    return {"items": items, "dispatch_rows": dispatch_rows}


def build_crm_client_index(api):
    crm_customers = api.request("GET", "inventory/crm/customers/")
    index = {}
    for c in crm_customers:
        name = str(c.get("nombre_cliente") or "").strip()
        if name:
            index[normalize_text(name)] = name
    return index


def map_dispatch_clients(dispatch_rows, crm_index):
    resolved = 0
    unresolved = 0
    for row in dispatch_rows:
        raw = row.get("cliente") or ""
        key = normalize_text(raw)
        match = crm_index.get(key)

        if not match and key:
            for crm_key, canonical in crm_index.items():
                if key in crm_key or crm_key in key:
                    match = canonical
                    break

        row["cliente_crm"] = match
        row["cliente_resuelto"] = bool(match)
        if match:
            resolved += 1
        else:
            unresolved += 1
    return {"resolved": resolved, "unresolved": unresolved}


def ensure_category_and_subcategories(api):
    category_name = "MIGRACION EXCEL DATACOM"
    target_sub = {
        "material": "MATERIAL",
        "herramienta": "HERRAMIENTA",
        "equipo": "EQUIPO",
        "general": "GENERAL",
    }

    categories = api.request("GET", "inventory/categories/")
    category_id = None
    for c in categories:
        if normalize_text(c.get("nombre_categoria")) == normalize_text(category_name):
            category_id = c["id"]
            break

    if not category_id:
        created = api.request(
            "POST",
            "inventory/categories/",
            data=json.dumps({"nombre_categoria": category_name}),
        )
        category_id = created["id"]

    all_sub = api.request("GET", "inventory/subcategories/")
    by_name = {}
    for s in all_sub:
        cat = s.get("categoria") or {}
        if str(cat.get("id", "")) == str(category_id):
            by_name[normalize_text(s.get("nombre"))] = s["id"]

    sub_map = {}
    for tipo, sub_name in target_sub.items():
        key = normalize_text(sub_name)
        if key not in by_name:
            created = api.request(
                "POST",
                "inventory/subcategories/",
                data=json.dumps({"nombre": sub_name, "categoria_id": category_id}),
            )
            by_name[key] = created["id"]
        sub_map[tipo] = by_name[key]

    return sub_map


def get_default_store_id(api):
    stores = api.request("GET", "inventory/stores/")
    preferred = ["bodega general conocoto", "mini bodega cumbaya"]
    normalized = {normalize_text(s.get("nombre_bodega")): s for s in stores}
    for key in preferred:
        if key in normalized:
            return normalized[key].get("id")
    if stores:
        return stores[0].get("id")
    return None


def upsert_items(api, items, subcategory_map, default_store_id, apply_changes):
    existing = api.request("GET", "inventory/items/")
    existing_by_code = {}
    if isinstance(existing, list):
        for i in existing:
            code = str(i.get("codigo") or "").strip()
            if code:
                existing_by_code[code] = i

    summary = {
        "created": 0,
        "updated": 0,
        "unchanged": 0,
        "errors": 0,
        "error_samples": [],
    }

    for item in items:
        code = item["codigo"]
        payload = {
            "codigo": code,
            "nombre": item["nombre"],
            "tipo_item": item["tipo_item"],
            "cantidad": int(item["cantidad"]),
            "subcategoria_id": subcategory_map[item["tipo_item"]],
        }

        if item["tipo_item"] == "equipo" and default_store_id:
            payload["ubicacion_actual_id"] = default_store_id

        current = existing_by_code.get(code)
        if current:
            changed = (
                str(current.get("nombre") or "") != payload["nombre"]
                or str(current.get("tipo_item") or "") != payload["tipo_item"]
                or int(current.get("cantidad") or 0) != payload["cantidad"]
            )
            if not changed:
                summary["unchanged"] += 1
                continue

            if apply_changes:
                try:
                    api.request("PUT", f"inventory/items/{current['id']}/", data=json.dumps(payload))
                    summary["updated"] += 1
                    time.sleep(0.2)
                except Exception as ex:
                    summary["errors"] += 1
                    if len(summary["error_samples"]) < 10:
                        summary["error_samples"].append({"codigo": code, "error": str(ex)})
            else:
                summary["updated"] += 1
        else:
            if apply_changes:
                try:
                    api.request("POST", "inventory/items/", data=json.dumps(payload))
                    summary["created"] += 1
                    time.sleep(0.2)
                except Exception as ex:
                    summary["errors"] += 1
                    if len(summary["error_samples"]) < 10:
                        summary["error_samples"].append({"codigo": code, "error": str(ex)})
            else:
                summary["created"] += 1

    return summary


def write_reports(output_dir, sync_summary, dispatch_rows):
    os.makedirs(output_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    summary_path = os.path.join(output_dir, f"inventory_sync_summary_{ts}.json")
    dispatch_path = os.path.join(output_dir, f"inventory_dispatch_report_{ts}.json")

    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(sync_summary, f, ensure_ascii=False, indent=2)

    with open(dispatch_path, "w", encoding="utf-8") as f:
        json.dump(dispatch_rows, f, ensure_ascii=False, indent=2)

    return {"summary": summary_path, "dispatch": dispatch_path}


def main():
    parser = argparse.ArgumentParser(description="Automatización de inventario desde Excel + CRM DataCom")
    parser.add_argument("--excel-path", required=True, help="Ruta al archivo .xlsm")
    parser.add_argument("--api-base", default="http://10.11.121.101:8070/api", help="Base URL API")
    parser.add_argument("--sso-token", required=True, help="SSO token ERP")
    parser.add_argument("--apply", action="store_true", help="Aplica cambios en API (por defecto dry-run)")
    parser.add_argument("--output-dir", default=".", help="Directorio para reportes JSON")
    args = parser.parse_args()

    api = ApiClient(base_url=args.api_base, sso_token=args.sso_token)
    api.login()

    excel_data = collect_excel_data(args.excel_path)
    crm_index = build_crm_client_index(api)
    crm_stats = map_dispatch_clients(excel_data["dispatch_rows"], crm_index)
    sub_map = ensure_category_and_subcategories(api)
    default_store_id = get_default_store_id(api)
    upsert_stats = upsert_items(
        api,
        excel_data["items"],
        sub_map,
        default_store_id,
        apply_changes=args.apply,
    )

    summary = {
        "mode": "apply" if args.apply else "dry-run",
        "excel_path": args.excel_path,
        "items_detected": len(excel_data["items"]),
        "dispatch_rows": len(excel_data["dispatch_rows"]),
        "crm_clients": len(crm_index),
        "crm_mapping": crm_stats,
        "upsert": upsert_stats,
        "default_store_id": default_store_id,
    }

    paths = write_reports(args.output_dir, summary, excel_data["dispatch_rows"])
    summary["reports"] = paths

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
